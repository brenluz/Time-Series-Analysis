"""
table_rmse_by_step.py
Produces: table_rmse_by_step.xlsx
Mean RMSE by model and forecast step (h=1..12),
averaged across all commodities and all Brazilian states.
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                               Border, Side, numbers)
from openpyxl.utils import get_column_letter

FILE   = "sliding_rmse_all_products.xlsx"
OUTPUT = "table_rmse_by_step.xlsx"
MODELS = ['ARIMA', 'ETS', 'Prophet', 'Random Forest',
          'LSTM', 'GRU', 'Transformer', 'Informer']

# ── 1. Aggregate ──────────────────────────────────────────────────────────────
xl          = pd.ExcelFile(FILE)
step_data   = {m: np.zeros(12) for m in MODELS}
step_counts = {m: np.zeros(12) for m in MODELS}

for sheet in xl.sheet_names:
    df        = pd.read_excel(FILE, sheet_name=sheet, header=None)
    model_row = df.iloc[0].tolist()
    data      = df.iloc[2:].copy()
    data.columns = range(data.shape[1])
    data      = data[data[0].notna()].reset_index(drop=True)

    for model in MODELS:
        cols = [i for i, v in enumerate(model_row) if v == model]
        for step_i, col_i in enumerate(cols):
            vals = pd.to_numeric(data[col_i], errors='coerce').dropna().values
            step_data[model][step_i]   += vals.sum()
            step_counts[model][step_i] += len(vals)

rows = []
for h in range(12):
    row = {m: round(step_data[m][h] / step_counts[m][h], 4) for m in MODELS}
    rows.append(row)

table_df = pd.DataFrame(rows, index=[f"h={i+1}" for i in range(12)])
table_df.index.name = "Forecast Step"

best_per_row = table_df.idxmin(axis=1)

# ── 2. Write xlsx ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "RMSE by Step"

# Colour palette
HEADER_BG   = "1A3A5C"   # dark navy
HEADER_FG   = "FFFFFF"
ROWLBL_BG   = "E8F0F7"   # light blue-grey
BEST_BG     = "1A6B3C"   # dark green  ← best cell background
BEST_FG     = "FFFFFF"
STRIPE_BG   = "F5F9FF"   # alternating row tint
WHITE_BG    = "FFFFFF"

thin  = Side(style='thin',   color='CCCCCC')
thick = Side(style='medium', color='1A3A5C')
cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
header_border = Border(left=thick, right=thick, top=thick, bottom=thick)

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left',   vertical='center')

# ── Header row ────────────────────────────────────────────────────────────────
ws.cell(row=1, column=1, value="Forecast Step").font = make_font(bold=True, color=HEADER_FG, size=10)
ws.cell(row=1, column=1).fill      = hfill(HEADER_BG)
ws.cell(row=1, column=1).alignment = center
ws.cell(row=1, column=1).border    = cell_border

for col_i, model in enumerate(MODELS, start=2):
    c = ws.cell(row=1, column=col_i, value=model)
    c.font      = make_font(bold=True, color=HEADER_FG, size=10)
    c.fill      = hfill(HEADER_BG)
    c.alignment = center
    c.border    = cell_border

# ── Data rows ─────────────────────────────────────────────────────────────────
for row_i, (step_label, series) in enumerate(table_df.iterrows(), start=2):
    bg = STRIPE_BG if row_i % 2 == 0 else WHITE_BG

    # Row label
    c = ws.cell(row=row_i, column=1, value=step_label)
    c.font      = make_font(bold=True, size=10)
    c.fill      = hfill(ROWLBL_BG)
    c.alignment = center
    c.border    = cell_border

    best_model = best_per_row[step_label]

    for col_i, model in enumerate(MODELS, start=2):
        val  = series[model]
        cell = ws.cell(row=row_i, column=col_i, value=val)
        cell.number_format = '0.0000'
        cell.alignment     = center
        cell.border        = cell_border

        if model == best_model:
            cell.fill = hfill(BEST_BG)
            cell.font = make_font(bold=True, color=BEST_FG, size=10)
        else:
            cell.fill = hfill(bg)
            cell.font = make_font(size=10)

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 16
for col_i in range(2, len(MODELS) + 2):
    ws.column_dimensions[get_column_letter(col_i)].width = 15

ws.row_dimensions[1].height = 20
for r in range(2, 14):
    ws.row_dimensions[r].height = 18

# ── Freeze panes & title ──────────────────────────────────────────────────────
ws.freeze_panes = "B2"

# Add a title above the table
ws.insert_rows(1)
title_cell = ws.cell(row=1, column=1,
    value="Table X — Mean RMSE by Model and Forecast Step")
title_cell.font      = make_font(bold=True, size=12, color="1A3A5C")
title_cell.alignment = left
ws.merge_cells(start_row=1, start_column=1,
               end_row=1,   end_column=len(MODELS) + 1)

# Add note below table
note_row = 15
ws.cell(row=note_row, column=1,
        value=("Note: Values are mean RMSE (BRL/kg) averaged across all 19 "
               "commodities and 26 Brazilian federative units. "
               "Green cells indicate the best-performing model at each forecast step."))
ws.cell(row=note_row, column=1).font      = make_font(size=8, color="555555")
ws.cell(row=note_row, column=1).alignment = left
ws.merge_cells(start_row=note_row, start_column=1,
               end_row=note_row,   end_column=len(MODELS) + 1)

ws.row_dimensions[1].height = 22

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")