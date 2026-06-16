"""
run_all_products.py
-------------------
Runs sliding_rmse_excel for every product sheet in the database and
writes all results to a single workbook: one tab per product.

Usage:
    python run_all_products.py

Output:
    sliding_rmse_all_products.xlsx
"""

import warnings
warnings.filterwarnings("ignore")

import os
import shutil
import logging
import argparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Project imports (all files must be in the same directory)
from analysis import sliding_rmse_excel

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
EXCEL_FILE_PATH = "../Databases/DatabaseConabv5.xlsx"
OUTPUT_XLSX     = "sliding_rmse_all_products.xlsx"
CACHE_DIR       = ".boxplot_cache"
TEST_PERIODS    = 12
MIN_TRAIN_SIZE  = 36
MAX_WINDOWS     = 12
EXCLUDE_STATES  = []   # e.g. ["RANDOM"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_sheet(ws):
    """Apply consistent professional formatting to a worksheet."""
    header_fill   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    subhead_fill  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
    subhead_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font     = Font(name="Arial", size=10)
    center        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left          = Alignment(horizontal="left",   vertical="center")
    thin          = Side(style="thin", color="BFBFBF")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row,
                                               min_col=1, max_col=max_col), start=1):
        for cell in row:
            cell.border = border
            if row_idx == 1:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center
            elif row_idx == 2:
                cell.fill      = subhead_fill
                cell.font      = subhead_font
                cell.alignment = center
            else:
                cell.font      = data_font
                if cell.column == 1:
                    cell.alignment = left
                else:
                    cell.alignment = center
                    # Zebra striping on data rows
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill("solid",
                                                start_color="EBF3FB",
                                                end_color="EBF3FB")
                    # Format numbers
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.00"

    # Auto-width for all columns
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len    = 0
        for cell in ws[col_letter]:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

    # Freeze panes below the two header rows, after the index column
    ws.freeze_panes = ws.cell(row=3, column=2)


def _flatten_multiindex_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Flatten a MultiIndex-column DataFrame into a two-row header DataFrame
    suitable for writing to Excel.

    MultiIndex columns like (Model, Step 1) become a plain DataFrame
    with a two-row header: first row = model names, second row = step labels.
    The index (states or windows) becomes the first column.
    """
    if not isinstance(df.columns, pd.MultiIndex):
        df.columns = [str(c) for c in df.columns]
        return df.reset_index()

    # Build two header rows from the MultiIndex levels
    level0 = [str(c[0]) for c in df.columns]   # Model names
    level1 = [str(c[1]) for c in df.columns]   # Step labels

    idx_name = df.index.name or "Index"
    header_row0 = [idx_name] + level0
    header_row1 = [""]       + level1

    data_rows = [[str(idx)] + [v for v in row]
                 for idx, row in zip(df.index, df.values)]

    flat = pd.DataFrame([header_row0, header_row1] + data_rows)
    return flat


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--max-cpu-cores", type=int, default=None)
    parser.add_argument("--device", choices=["cuda", "cpu", "auto"], default="auto")
    args = parser.parse_args()

    os.environ["ML_DEVICE"] = args.device
    if args.max_cpu_cores:
        os.environ["ML_MAX_CPU_CORES"] = str(args.max_cpu_cores)
    n_jobs = args.max_cpu_cores if args.max_cpu_cores else -1

    # --- Clear cache ---
    if os.path.exists(CACHE_DIR):
        logger.info(f"Clearing cache at {CACHE_DIR} …")
        shutil.rmtree(CACHE_DIR)
    os.makedirs(CACHE_DIR, exist_ok=True)

    # --- Discover all product sheets ---
    xl         = pd.ExcelFile(EXCEL_FILE_PATH)
    all_sheets = xl.sheet_names
    logger.info(f"Found {len(all_sheets)} product sheets: {all_sheets}")

    # --- Temporary per-product xlsx files ---
    tmp_files = {}
    for sheet_name in all_sheets:
        logger.info(f"\n{'='*60}\nProcessing product: {sheet_name}\n{'='*60}")
        try:
            df = pd.read_excel(EXCEL_FILE_PATH, sheet_name=sheet_name, index_col=0)
            df.index.name = "Date"
            if EXCLUDE_STATES:
                df = df.drop(columns=EXCLUDE_STATES, errors="ignore")

            tmp_path = f"_tmp_{sheet_name}.xlsx"
            sliding_rmse_excel(
                df,
                product_name=sheet_name,
                test_periods=TEST_PERIODS,
                min_train_size=MIN_TRAIN_SIZE,
                max_windows=MAX_WINDOWS,
                output_xlsx=tmp_path,
                cache_dir=CACHE_DIR,
                n_jobs=n_jobs,
            )
            tmp_files[sheet_name] = tmp_path
            logger.info(f"Done: {sheet_name} -> {tmp_path}")

        except Exception as exc:
            logger.error(f"Failed on {sheet_name}: {exc}")

    # --- Merge all tmp files into one workbook ---
    logger.info("\nMerging all products into single workbook …")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for sheet_name, tmp_path in tmp_files.items():
            if not os.path.exists(tmp_path):
                continue
            try:
                # Read the Summary sheet from each tmp file
                summary_df = pd.read_excel(tmp_path, sheet_name="Summary",
                                           index_col=0, header=[0, 1])
                tab_name   = str(sheet_name)[:31]

                flat = _flatten_multiindex_df(summary_df)
                flat.to_excel(writer, sheet_name=tab_name,
                              index=False, header=False)
            except Exception as exc:
                logger.error(f"Could not merge {sheet_name}: {exc}")

    # --- Apply formatting ---
    wb = load_workbook(OUTPUT_XLSX)
    for ws in wb.worksheets:
        _style_sheet(ws)
    wb.save(OUTPUT_XLSX)

    # --- Clean up tmp files ---
    for tmp_path in tmp_files.values():
        try:
            os.remove(tmp_path)
        except Exception:
            pass

    logger.info(f"\nAll done. Output: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
    # All output files are written and flushed inside main(). Some dependencies
    # (kaleido's Chrome subprocess, torch thread pools, the Windows process
    # pool) can leave background threads/subprocesses alive that hang the
    # interpreter at shutdown and freeze the terminal. Force an immediate,
    # clean exit so control returns to the shell.
    import sys
    sys.stdout.flush()
    sys.stderr.flush()
    os._exit(0)