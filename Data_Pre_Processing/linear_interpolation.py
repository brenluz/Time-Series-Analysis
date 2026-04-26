"""
CONAB Interpolation Script — Corrected Version
================================================
    .interpolate(method='linear', limit_direction='forward')

    'forward' only fills NaN values that have a valid observation on BOTH
    sides (i.e., genuine interior gaps). Leading NaN blocks (no left anchor)
    and trailing NaN blocks (no right anchor) are left as NaN and the
    corresponding state-commodity pairs are excluded from that sheet.

The script writes a clean Excel file and a detailed audit CSV so you can
inspect exactly which state-commodity pairs were dropped and why.

Usage:
    python interpolate_conab.py
    python interpolate_conab.py --input path/to/DatabaseConabv4.xlsx \
                                --output path/to/DatabaseConabv5_corrected.xlsx \
                                --audit  path/to/audit.csv
"""

import argparse
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Default paths ─────────────────────────────────────────────────────────────

INPUT_FILE  = '../Databases/DatabaseConabv4.xlsx'
OUTPUT_FILE = '../Databases/DatabaseConabv5_corrected.xlsx'
AUDIT_FILE  = '../Databases/interpolation_audit.csv'


# ── Core logic ────────────────────────────────────────────────────────────────

def has_leading_nans(series: pd.Series) -> bool:
    """True if the series begins with one or more NaN values."""
    return pd.isna(series.iloc[0])


def has_trailing_nans(series: pd.Series) -> bool:
    """True if the series ends with one or more NaN values."""
    return pd.isna(series.iloc[-1])


def count_leading_nans(series: pd.Series) -> int:
    count = 0
    for v in series:
        if pd.isna(v):
            count += 1
        else:
            break
    return count


def count_trailing_nans(series: pd.Series) -> int:
    count = 0
    for v in reversed(series.values):
        if pd.isna(v):
            count += 1
        else:
            break
    return count


def process_sheet(sheet_name: str, df_raw: pd.DataFrame,
                  audit_rows: list) -> pd.DataFrame:
    """
    Process one sheet:
      1. Convert all state columns to numeric.
      2. Drop fully-missing columns (Type 4).
      3. Drop columns with leading or trailing NaNs (cannot interpolate safely).
      4. Interpolate interior gaps only (limit_direction='forward' with both
         endpoints guaranteed non-NaN after step 3).

    Returns a cleaned DataFrame with only valid, fully-interpolated columns.
    """
    df = df_raw.copy()
    state_cols = [c for c in df.columns if c != 'Date']

    for col in state_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    kept_cols   = []
    dropped_cols = []

    for col in state_cols:
        series = df[col]
        n_total   = len(series)
        n_missing = int(series.isna().sum())
        n_valid   = n_total - n_missing

        # ── Fully missing ──────────────────────────────────────────────────
        if n_valid == 0:
            audit_rows.append({
                'sheet': sheet_name, 'state': col,
                'action': 'DROPPED',
                'reason': 'fully missing (0 valid observations)',
                'n_valid': 0, 'n_missing': n_missing,
                'n_leading_nan': n_missing, 'n_trailing_nan': 0,
                'n_interior_gaps_filled': 0,
            })
            dropped_cols.append(col)
            continue

        n_lead  = count_leading_nans(series)
        n_trail = count_trailing_nans(series)

        # ── Leading or trailing NaNs ───────────────────────────────────────
        if n_lead > 0 or n_trail > 0:
            reason_parts = []
            if n_lead  > 0: reason_parts.append(f'{n_lead} leading NaN(s)')
            if n_trail > 0: reason_parts.append(f'{n_trail} trailing NaN(s)')
            audit_rows.append({
                'sheet': sheet_name, 'state': col,
                'action': 'DROPPED',
                'reason': '; '.join(reason_parts) + ' — cannot interpolate without anchor',
                'n_valid': n_valid, 'n_missing': n_missing,
                'n_leading_nan': n_lead, 'n_trailing_nan': n_trail,
                'n_interior_gaps_filled': 0,
            })
            dropped_cols.append(col)
            continue

        # ── Interior gaps only — safe to interpolate ───────────────────────
        n_interior = int(series.isna().sum())  # leading/trailing already 0 here
        if n_interior > 0:
            df[col] = series.interpolate(method='linear', limit_direction='forward')
            audit_rows.append({
                'sheet': sheet_name, 'state': col,
                'action': 'INTERPOLATED',
                'reason': f'{n_interior} interior gap(s) filled via linear interpolation',
                'n_valid': n_valid, 'n_missing': n_missing,
                'n_leading_nan': 0, 'n_trailing_nan': 0,
                'n_interior_gaps_filled': n_interior,
            })
        else:
            audit_rows.append({
                'sheet': sheet_name, 'state': col,
                'action': 'KEPT_AS_IS',
                'reason': 'no missing values',
                'n_valid': n_valid, 'n_missing': 0,
                'n_leading_nan': 0, 'n_trailing_nan': 0,
                'n_interior_gaps_filled': 0,
            })
        kept_cols.append(col)

    # Return only kept columns
    result = df[['Date'] + kept_cols].copy()

    # Final NaN check — should be zero after correct interpolation
    remaining = result[kept_cols].isna().sum().sum()
    if remaining > 0:
        print(f'  [WARN] {sheet_name}: {remaining} NaN(s) remain after '
              f'interpolation — check audit log.')

    return result


# ── Writer ────────────────────────────────────────────────────────────────────

def write_excel(cleaned: dict[str, pd.DataFrame], output_path: str) -> None:
    HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
    DATE_FILL    = PatternFill('solid', fgColor='2E75B6')
    HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    DATA_FONT    = Font(name='Arial', size=10)
    DATE_FONT    = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    BORDER_SIDE  = Side(style='thin', color='BFBFBF')
    THIN_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                          top=BORDER_SIDE, bottom=BORDER_SIDE)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet, df in cleaned.items():
        ws = wb.create_sheet(title=sheet)
        cols = df.columns.tolist()

        # Header
        for c_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = THIN_BORDER

        # Data
        for r_idx, row in df.iterrows():
            excel_row = r_idx + 2
            for c_idx, col_name in enumerate(cols, start=1):
                val = row[col_name]
                if isinstance(val, (np.floating, np.float64)):
                    val = float(val) if not np.isnan(val) else None
                elif isinstance(val, np.integer):
                    val = int(val)
                cell = ws.cell(row=excel_row, column=c_idx, value=val)
                cell.border = THIN_BORDER
                if c_idx == 1:
                    cell.font      = DATE_FONT
                    cell.fill      = DATE_FILL
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.font      = DATA_FONT
                    if val is not None:
                        cell.number_format = '0.000'
                    cell.alignment = Alignment(horizontal='right')

        ws.column_dimensions[get_column_letter(1)].width = 12
        for c_idx in range(2, len(cols) + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 8
        ws.freeze_panes = 'B2'

    wb.save(output_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def run(input_path: str, output_path: str, audit_path: str) -> None:
    print(f'Loading: {input_path}')
    raw_sheets = pd.read_excel(input_path, sheet_name=None, index_col=0)
    sheets = list(raw_sheets.keys())
    print(f'Sheets: {sheets}\n')

    audit_rows: list[dict] = []
    cleaned:    dict[str, pd.DataFrame] = {}

    for sheet_name, df_raw in raw_sheets.items():
        # Reset index so Date becomes a regular column
        df = df_raw.reset_index()
        df = df.rename(columns={df.columns[0]: 'Date'})
        df = df[df['Date'].notna()].reset_index(drop=True)

        print(f'Processing: {sheet_name}  ({len([c for c in df.columns if c != "Date"])} states)')
        result = process_sheet(sheet_name, df, audit_rows)
        kept = len([c for c in result.columns if c != 'Date'])
        print(f'  → {kept} states retained\n')
        cleaned[sheet_name] = result

    # Write cleaned Excel
    print(f'Writing: {output_path}')
    write_excel(cleaned, output_path)
    print('Done.\n')

    # Write audit CSV
    audit_df = pd.DataFrame(audit_rows)
    audit_df.to_csv(audit_path, index=False)
    print(f'Audit written: {audit_path}\n')

    # Print summary
    print('=' * 65)
    print(f'{"Sheet":<30} {"Kept":>6} {"Dropped":>8} {"Interpolated":>13}')
    print('-' * 65)
    for sheet in sheets:
        sheet_audit = audit_df[audit_df['sheet'] == sheet]
        kept   = (sheet_audit['action'].isin(['KEPT_AS_IS', 'INTERPOLATED'])).sum()
        drop   = (sheet_audit['action'] == 'DROPPED').sum()
        interp = (sheet_audit['action'] == 'INTERPOLATED').sum()
        print(f'{sheet:<30} {kept:>6} {drop:>8} {interp:>13}')
    print('=' * 65)
    total_dropped = (audit_df['action'] == 'DROPPED').sum()
    total_interp  = (audit_df['action'] == 'INTERPOLATED').sum()
    print(f'{"TOTAL":<30} {"":>6} {total_dropped:>8} {total_interp:>13}')
    print()
    print('Dropped state-commodity pairs:')
    dropped = audit_df[audit_df['action'] == 'DROPPED'][['sheet','state','reason']]
    for _, row in dropped.iterrows():
        print(f'  {row["sheet"]:<30} {row["state"]}  —  {row["reason"]}')


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Correct interpolation of CONAB database (fixes limit_direction bug).')
    parser.add_argument('--input',  default=INPUT_FILE)
    parser.add_argument('--output', default=OUTPUT_FILE)
    parser.add_argument('--audit',  default=AUDIT_FILE)
    args = parser.parse_args()
    run(args.input, args.output, args.audit)