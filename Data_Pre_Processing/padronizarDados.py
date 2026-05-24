import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INPUT_FILE = "../Databases/Consulta_2025.xlsx"
OUTPUT_FILE = "Consulta_2025_pivotada.xlsx"

# Load and forward-fill merged cells
df = pd.read_excel(INPUT_FILE)
df = df.ffill()

# Format the date column as "Jan/2025" style strings
df["Período"] = pd.to_datetime(df["Período"]).dt.strftime("%b/%Y")

# Pivot: rows = (UF, Período), columns = Produto, values = Preço medio
pivot = df.pivot_table(
    index=["UF", "Período"],
    columns="Produto",
    values="Preço medio",
    aggfunc="mean"
)

# Reset index so UF and Período become regular columns
pivot = pivot.reset_index()
pivot.columns.name = None

# Reorder rows by UF alphabetically and Período chronologically
month_order = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
pivot["_month_sort"] = pivot["Período"].str[:3].map({m: i for i, m in enumerate(month_order)})
pivot = pivot.sort_values(["UF", "_month_sort"]).drop(columns="_month_sort").reset_index(drop=True)

# Save to Excel
pivot.to_excel(OUTPUT_FILE, index=False)

# ── Formatting ──────────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

header_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
subhdr_fill  = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
alt_fill     = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
white_fill   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
thin_side    = Side(style="thin", color="BDD7EE")
thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Header row
for col_idx, cell in enumerate(ws[1], start=1):
    if col_idx <= 2:
        cell.fill = subhdr_fill
    else:
        cell.fill = header_fill
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

ws.row_dimensions[1].height = 45

# Data rows
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    for col_idx, cell in enumerate(row, start=1):
        cell.fill = fill
        cell.border = thin_border
        if col_idx <= 2:
            cell.font = Font(name="Arial", bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value is not None:
                cell.number_format = "#,##0.00"

# Column widths
ws.column_dimensions["A"].width = 6   # UF
ws.column_dimensions["B"].width = 12  # Período
for col in ws.iter_cols(min_col=3, max_col=ws.max_column):
    letter = col[0].column_letter
    ws.column_dimensions[letter].width = 14

# Freeze panes (keep UF + Período + header visible)
ws.freeze_panes = "C2"

wb.save(OUTPUT_FILE)
print(f"Done! Saved to {OUTPUT_FILE}")
print(f"Shape: {pivot.shape[0]} rows x {pivot.shape[1]} columns")